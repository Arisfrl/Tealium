import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.max_row


def getColumnCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.max_column


def readData(file: object, sheetName: object, row: object, col: object) -> object:
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    return ws.cell(row, col).value


def writeData(file, sheetName, row, col, data):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(row, col).value = data
    wb.save(file)


def fillGreenColor(file, sheetName, row, col):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(row, col).fill = PatternFill("solid", fgColor="00FF00")
    wb.save(file)


def fillRedColor(file, sheetName, row, col):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(row, col).fill = PatternFill("solid", fgColor="FF0000")
    wb.save(file)
